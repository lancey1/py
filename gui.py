import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import scrolledtext


def process_files(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(os.path.join(folder_path, filename), data_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                output_box.insert(tk.END, f"Sheet: {sheetname}\n")
            for row in sheet.iter_rows(values_only=True):
                row_values = []
                for value in row:
                    if value is not None:
                        row_values.append(value)
                if row_values:
                    output_box.insert(tk.END, f"{row_values}\n")


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.geometry('400x400')
        self.master.title('File Explorer')
        self.folder_path = tk.StringVar()
        self.create_widgets()

    def create_widgets(self):
        self.select_folder_button = tk.Button(self.master, text="Select Folder", command=self.select_folder)
        self.select_folder_button.pack(pady=10)

        self.run_button = tk.Button(self.master, text="Run", command=self.run_program)
        self.run_button.pack(pady=10)

        global output_box
        output_box = scrolledtext.ScrolledText(self.master, height=10, width=50)
        output_box.pack(pady=10, fill=tk.BOTH, expand=True)

        self.clear_button = tk.Button(self.master, text="Clear", command=self.clear_output)
        self.clear_button.pack(pady=10)

    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        self.folder_path.set(folder_selected)

    def run_program(self):
        folder_path = self.folder_path.get()
        if folder_path:
            process_files(folder_path)

    def clear_output(self):
        output_box.delete('1.0', tk.END)


root = tk.Tk()
app = Application(master=root)
app.mainloop()